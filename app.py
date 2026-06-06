import streamlit as st
import pandas as pd
from PIL import Image
import os
import base64
import io
import hmac
import random
from xml.sax.saxutils import escape
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# ১. পেজ কনফিগারেশন সেটআপ
st.set_page_config(
    page_title="Exam Duty Portal | Narsingdi GPI", 
    page_icon="📝", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# সেশন স্টেট ইনিশিয়ালাইজেশন
if "assignments" not in st.session_state:
    st.session_state["assignments"] = {}

if "language" not in st.session_state:
    st.session_state["language"] = "bn"

if "custom_rooms" not in st.session_state:
    st.session_state["custom_rooms"] = [f"Room {i}" for i in range(101, 110)]  # ডিফল্ট ৯টি রুম

if "custom_floors" not in st.session_state:
    st.session_state["custom_floors"] = ["Ground Floor", "1st Floor", "2nd Floor", "3rd Floor", "4th Floor"]  # ডিফল্ট ৫টি ফ্লোর

# লোগো ফাইল অটো-চেক করার ফাংশন
def get_logo_base64():
    possible_names = ["NPI Logo.jpg", "NPI Logo.JPG", "NPI Logo.jpeg", "NPI Logo.png", "npi logo.jpg"]
    for name in possible_names:
        if os.path.exists(name):
            try:
                with open(name, "rb") as img_file:
                    encoded_string = base64.b64encode(img_file.read()).decode()
                    ext = name.split('.')[-1].lower()
                    mime_type = "png" if ext == "png" else "jpeg"
                    return f"data:image/{mime_type};base64,{encoded_string}"
            except Exception:
                pass
    return None

logo_b64 = get_logo_base64()

# ২. মডার্ন ইউজার ইন্টারফেস ডিজাইন (CSS) - লাইন স্পেস ও প্যাডিং মিনিমাইজ করা হয়েছে
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Hind+Siliguri:wght=400;500;600;700&family=Inter:wght=400;500;600;700&display=swap');
    
    html, body, .stSelectbox, div[data-testid="stMarkdownContainer"] {
        font-family: 'Hind Siliguri', 'Inter', sans-serif;
    }
    
    .portal-header-container {
        display: flex;
        align-items: center;
        background: linear-gradient(135deg, #1E3A8A 0%, #0F172A 100%);
        padding: 20px 30px;
        border-radius: 12px;
        margin-bottom: 20px;
        box-shadow: 0 4px 15px rgba(30, 58, 138, 0.1);
        border-bottom: 4px solid #F59E0B;
    }
    .header-img-render {
        width: 75px;
        height: 75px;
        margin-right: 20px;
        border-radius: 8px;
        background-color: white;
        padding: 2px;
    }
    .fallback-logo-render {
        font-size: 45px;
        margin-right: 20px;
        background: white;
        padding: 10px;
        border-radius: 8px;
        line-height: 1;
    }
    .header-text h1 {
        font-size: 26px !important;
        font-weight: 700 !important;
        margin: 0 0 4px 0 !important;
        color: #FFFFFF !important;
    }
    .header-text p {
        font-size: 15px;
        color: #CBD5E1;
        margin: 0;
    }
    
    .profile-card {
        background: #F8FAFC;
        border: 1px solid #E2E8F0;
        border-left: 5px solid #1E3A8A;
        padding: 18px;
        border-radius: 8px;
        margin-top: 20px;
        margin-bottom: 25px;
    }
    .profile-grid {
        display: grid;
        grid-template-columns: 1fr;
        gap: 12px;
    }
    @media (min-width: 768px) {
        .profile-grid { grid-template-columns: 1fr 1fr; gap: 20px; }
    }
    .profile-item { font-size: 15px; color: #475569; }
    .profile-item strong { color: #0F172A; font-weight: 600; display: inline-block; width: 110px; }

    /* গ্রিডের লাইন স্পেস এবং এলিমেন্টের মার্জিন সর্বনিম্ন করার কাস্টম CSS */
    div[data-testid="stVerticalBlock"] > div {
        gap: 0rem !important; /* ব্লকগুলোর ভেতরের গ্যাপ কমানো */
    }
    
    /* স্ট্রিমলিট চেকবক্সের ফন্ট সাইজ এবং লাইন গ্যাপ একদম মিনিমাইজ করা */
    div[data-testid="stCheckbox"] {
        margin-bottom: 2px !important;
        margin-top: 2px !important;
        padding: 0px !important;
    }
    div[data-testid="stCheckbox"] label p {
        font-size: 12.5px !important;
        font-weight: 500 !important;
        line-height: 1.1 !important;
    }

    /* অফিসিয়াল ডাউনলোড বোতামের কাস্টমাইজেশন */
    div.stDownloadButton {
        text-align: center;
        margin: 30px 0;
    }
    div.stDownloadButton > button {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        color: white !important;
        padding: 14px 50px !important;
        border: none !important;
        border-radius: 6px !important;
        font-size: 16px !important;
        font-weight: 600 !important;
        width: 100% !important;
        box-shadow: 0 4px 12px rgba(16, 185, 129, 0.3) !important;
        transition: all 0.2s ease !important;
    }
    div.stDownloadButton > button:hover {
        box-shadow: 0 6px 20px rgba(16, 185, 129, 0.4) !important;
        transform: translateY(-1px) !important;
    }
    </style>
""", unsafe_allow_html=True)

# ৩. ডেটা প্রসেসিং ইঞ্জিন
@st.cache_data
def load_and_process_roster():
    file_path = "Duty Roaster for Mid Term.xlsx"
    sheets_config = {
        "Chief Invigilator": "All Chief Invigilators",
        "Hall Super": "All Hallsupers",
        "Teacher / Invigilator": "All Teachers",
        "MLSS / Staff": "MLSS Roaster"
    }
    
    parsed_records = []
    
    for user_role, sheet_name in sheets_config.items():
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            dates_row = df.iloc[1].ffill() 
            times_row = df.iloc[2]
            
            for idx in range(3, len(df)):
                row = df.iloc[idx]
                if pd.isna(row[1]) or str(row[1]).strip() == "" or "Total" in str(row[1]):
                    continue
                    
                name = str(row[1]).strip()
                designation = str(row[2]).strip() if pd.notna(row[2]) else "N/A"
                text_dept = str(row[3]).strip() if pd.notna(row[3]) else "N/A"
                
                for col_idx in range(4, len(row) - 1):
                    cell_value = row[col_idx]
                    if pd.notna(cell_value) and str(cell_value).strip() != "" and cell_value != 0:
                        date_str = str(dates_row[col_idx]).split(" ")[0]
                        time_slot = str(times_row[col_idx]).strip()
                        
                        if time_slot in ["9.3", "9.30"]: time_slot = "09:30 AM"
                        elif time_slot in ["11.0", "11"]: time_slot = "11:00 AM"
                        elif time_slot in ["12.3", "12.30"]: time_slot = "12:30 PM"
                        elif time_slot in ["2.3", "2.30"]: time_slot = "02:30 PM"
                        
                        clean_val = str(cell_value).strip().replace('.0','')
                        
                        if clean_val in ["1", "1.0", "Assigned"]:
                            duty_desc = "Assigned"
                        elif user_role == "MLSS / Staff":
                            duty_desc = f"Shift: {clean_val}" if clean_val.isdigit() else clean_val
                        else:
                            duty_desc = f"Room: {clean_val}"
                            
                        parsed_records.append({
                            "Role": user_role,
                            "Name": name,
                            "Designation": designation,
                            "Technology/Dept": text_dept,
                            "Date": date_str,
                            "Time Slot": time_slot,
                            "Duty Status": duty_desc
                        })
        except Exception as e:
            st.warning(f"Error handling sheet {sheet_name}: {e}")
            
    return pd.DataFrame(parsed_records)

# ৩.৫ আসল পিডিএফ জেনারেটর (ReportLab)
def generate_pdf(user_info, schedule_df):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm,
        title=f"Duty Roster - {user_info['Name']}",
    )

    navy = colors.HexColor("#1E3A8A")
    slate = colors.HexColor("#475569")
    border = colors.HexColor("#CBD5E1")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DocTitle", parent=styles["Title"],
                                 textColor=navy, fontSize=18, spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle("DocSub", parent=styles["Normal"],
                               textColor=slate, fontSize=11, alignment=TA_CENTER, spaceAfter=16)
    cell_style = ParagraphStyle("ProfileCell", parent=styles["Normal"],
                                fontSize=10.5, textColor=colors.HexColor("#0F172A"))

    def profile_cell(label, value):
        return Paragraph(f"<b>{escape(str(label))}:</b> {escape(str(value))}", cell_style)

    elements = [
        Paragraph("Narsingdi Government Polytechnic Institute", title_style),
        Paragraph("Mid-Term Examination - 2026  |  Personalized Duty Roster", sub_style),
    ]

    profile = Table(
        [[profile_cell("Name", user_info["Name"]), profile_cell("Dept/Tech", user_info["Technology/Dept"])],
         [profile_cell("Designation", user_info["Designation"]), profile_cell("Category", user_info["Role"])]],
        colWidths=[doc.width / 2.0] * 2,
    )
    profile.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.75, border),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements += [profile, Spacer(1, 16)]

    data = [["SL", "Date", "Time Slot", "Duty Status", "Floor/Room"]]
    for i, (_, row) in enumerate(schedule_df.iterrows(), start=1):
        f_r = row.get("Floor / Room Assignment", "N/A")
        data.append([str(i), str(row["Date"]), str(row["Time Slot"]), str(row["Duty Status"]), str(f_r)])
    table = Table(data, colWidths=[doc.width * w for w in (0.08, 0.22, 0.22, 0.24, 0.24)], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, border),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

# ৩.৬ কন্ট্রোল রুম ড্যাশবোর্ড — পাসওয়ার্ড সুরক্ষিত
_DEFAULT_PASSWORD = "ngpi"
ADMIN_USER = os.environ.get("CONTROL_ROOM_USER", "CRO")
ADMIN_PASSWORD = os.environ.get("CONTROL_ROOM_PASSWORD", _DEFAULT_PASSWORD)


def _check_admin_login():
    if st.session_state.get("cr_authenticated"):
        return True

    st.markdown("#### 🔒 Control Room Login")
    st.caption("Restricted area — authorised personnel only.")
    with st.form("control_room_login", clear_on_submit=False):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in")

    if submitted:
        ok_user = hmac.compare_digest(username.strip(), ADMIN_USER)
        ok_pass = hmac.compare_digest(password, ADMIN_PASSWORD)
        if ok_user and ok_pass:
            st.session_state["cr_authenticated"] = True
            st.rerun()
        else:
            st.error("❌ Invalid username or password.")
    return False


def run_random_assignment(df, active_rooms, active_floors):
    """Algorithmic allocator mapped strictly by role: Teachers -> Room, Hall Super/MLSS -> Floor"""
    new_assignments = {}
    grouped = df.groupby(['Date', 'Time Slot'])
    
    for (date, slot), group in grouped:
        shift_key = f"{date}_{slot}"
        new_assignments[shift_key] = {}
        
        # ১. Teacher / Invigilator -> ROOM অ্যাসাইনমেন্ট
        invigilators = group[group['Role'] == 'Teacher / Invigilator']['Name'].tolist()
        if invigilators and active_rooms:
            random.shuffle(invigilators)
            room_slots = [[r, 1] for r in active_rooms] 
            
            idx = 0
            while len(invigilators) > len(room_slots) and idx < len(room_slots):
                room_slots[idx][1] = 2
                idx += 1
                
            current_inv_idx = 0
            for r_num, count in room_slots:
                for _ in range(count):
                    if current_inv_idx < len(invigilators):
                        new_assignments[shift_key][invigilators[current_inv_idx]] = f"Room: {r_num}"
                        current_inv_idx += 1
                        
            while current_inv_idx < len(invigilators):
                r_rand = random.choice(active_rooms)
                new_assignments[shift_key][invigilators[current_inv_idx]] = f"Room: {r_rand}"
                current_inv_idx += 1

        # ২. Hall Super -> FLOOR অ্যাসাইনমেন্ট
        hall_supers = group[group['Role'] == 'Hall Super']['Name'].tolist()
        if hall_supers and active_floors:
            random.shuffle(hall_supers)
            floor_slots = [[f, 1] for f in active_floors]
            
            idx = 0
            while len(hall_supers) > len(floor_slots) and idx < len(floor_slots):
                floor_slots[idx][1] = 2
                idx += 1
                
            current_hs_idx = 0
            for f_num, count in floor_slots:
                for _ in range(count):
                    if current_hs_idx < len(hall_supers):
                        new_assignments[shift_key][hall_supers[current_hs_idx]] = f"{f_num}"
                        current_hs_idx += 1
            while current_hs_idx < len(hall_supers):
                f_rand = random.choice(active_floors)
                new_assignments[shift_key][hall_supers[current_hs_idx]] = f"{f_rand}"
                current_hs_idx += 1

        # ৩. MLSS / Staff -> FLOOR অ্যাসাইনমেন্ট
        mlss_staff = group[group['Role'] == 'MLSS / Staff']['Name'].tolist()
        if mlss_staff and active_floors:
            random.shuffle(mlss_staff)
            mlss_slots = [[f, 1] for f in active_floors]
            
            idx = 0
            while len(mlss_staff) > len(mlss_slots) and idx < len(mlss_slots):
                mlss_slots[idx][1] = 2
                idx += 1
                
            current_m_idx = 0
            for f_num, count in mlss_slots:
                for _ in range(count):
                    if current_m_idx < len(mlss_staff):
                        new_assignments[shift_key][mlss_staff[current_m_idx]] = f"{f_num}"
                        current_m_idx += 1
            while current_m_idx < len(mlss_staff):
                f_rand = random.choice(active_floors)
                new_assignments[shift_key][mlss_staff[current_m_idx]] = f"{f_rand}"
                current_m_idx += 1

    st.session_state["assignments"] = new_assignments
    st.success("🎯 Floor & Room configuration generated matching exact institutional role requirements!")


def render_control_room(df):
    st.markdown("## 🛡️ Control Room Dashboard")

    if not _check_admin_login():
        return

    head = st.columns([3, 1])
    head[0].success(f"Logged in as **{ADMIN_USER}**")
    if head[1].button("🔓 Log out", use_container_width=True):
        st.session_state["cr_authenticated"] = False
        st.rerun()

    if ADMIN_PASSWORD == _DEFAULT_PASSWORD:
        st.warning(
            "Using the default password. Set the `CONTROL_ROOM_PASSWORD` "
            "environment variable to secure this dashboard.",
            icon="⚠️",
        )

    # ⚙️ Advanced Assignment Settings Panel
    with st.expander("⚙️ Advanced Floor & Room Assignment Matrix", expanded=True):
        st.markdown("##### Setup active rooms/floors for duty configuration:")
        
        c_conf1, c_conf2 = st.columns(2)
        
        # --- ১. রুম ম্যানেজমেন্ট (Teachers) ---
        with c_conf1:
            st.markdown("#### 🚪 Room Settings (For Teachers)")
            
            # নতুন রুম যোগ করা
            with st.container():
                r_add_col1, r_add_col2 = st.columns([3, 1])
                with r_add_col1:
                    new_room_input = st.text_input("➕ Add New Room:", placeholder="e.g., Room 205, 301", key="txt_add_room", label_visibility="collapsed")
                with r_add_col2:
                    if st.button("Add", key="btn_add_room", use_container_width=True) and new_room_input.strip():
                        r_clean = new_room_input.strip()
                        if r_clean.isdigit():
                            r_clean = f"Room {r_clean}"
                        if r_clean not in st.session_state["custom_rooms"]:
                            st.session_state["custom_rooms"].append(r_clean)
                            st.session_state["custom_rooms"].sort()
                            st.rerun()
            
            # রুমের নাম পরিবর্তন (Rename Section)
            if st.session_state["custom_rooms"]:
                with st.container():
                    r_ren_col1, r_ren_col2, r_ren_col3 = st.columns([1.5, 2, 1])
                    with r_ren_col1:
                        target_room = st.selectbox("Select to Rename:", st.session_state["custom_rooms"], key="sb_target_room", label_visibility="collapsed")
                    with r_ren_col2:
                        renamed_room = st.text_input("New Name:", placeholder="New room name", key="txt_rename_room", label_visibility="collapsed")
                    with r_ren_col3:
                        if st.button("Rename", key="btn_rename_room", use_container_width=True) and renamed_room.strip():
                            r_new_clean = renamed_room.strip()
                            if r_new_clean.isdigit():
                                r_new_clean = f"Room {r_new_clean}"
                            
                            idx = st.session_state["custom_rooms"].index(target_room)
                            st.session_state["custom_rooms"][idx] = r_new_clean
                            st.session_state["custom_rooms"].sort()
                            st.rerun()

            # রুম ডিলিট করার ডেডিকেটেড অপশন (Delete Section)
            if st.session_state["custom_rooms"]:
                with st.container():
                    r_del_col1, r_del_col2 = st.columns([3.5, 1])
                    with r_del_col1:
                        delete_target_room = st.selectbox("Select to Delete:", st.session_state["custom_rooms"], key="sb_delete_room", label_visibility="collapsed")
                    with r_del_col2:
                        if st.button("Delete", key="btn_delete_room", use_container_width=True):
                            st.session_state["custom_rooms"].remove(delete_target_room)
                            st.rerun()

            st.markdown("**Active Rooms (Check to activate):**")
            selected_rooms = []
            
            # 🚀 ৫ কলাম বিশিষ্ট কাস্টম স্পেস-সেভিং গ্রিড লেআউট (লাইন স্পেস মিনিমাইজড)
            if st.session_state["custom_rooms"]:
                room_list = list(st.session_state["custom_rooms"])
                cols_per_row = 5
                for i in range(0, len(room_list), cols_per_row):
                    row_rooms = room_list[i:i+cols_per_row]
                    grid_cols = st.columns(cols_per_row)
                    for idx, r in enumerate(row_rooms):
                        with grid_cols[idx]:
                            if st.checkbox(r, value=True, key=f"chk_{r}"):
                                val_to_append = r.split(" ")[1] if " " in r else r
                                selected_rooms.append(val_to_append)
                        
        # --- ২. ফ্লোর ম্যানেজমেন্ট (Hall Supers & MLSS) ---
        with c_conf2:
            st.markdown("#### 🏢 Floor Settings (For Hall Supers & MLSS)")
            
            # নতুন ফ্লোর যোগ করা
            with st.container():
                f_add_col1, f_add_col2 = st.columns([3, 1])
                with f_add_col1:
                    new_floor_input = st.text_input("➕ Add New Floor:", placeholder="e.g., 5th Floor, Ground Floor", key="txt_add_floor", label_visibility="collapsed")
                with f_add_col2:
                    if st.button("Add", key="btn_add_floor", use_container_width=True) and new_floor_input.strip():
                        f_clean = new_floor_input.strip()
                        if f_clean not in st.session_state["custom_floors"]:
                            st.session_state["custom_floors"].append(f_clean)
                            st.rerun()
            
            # ফ্লোরের নাম পরিবর্তন (Rename Section)
            if st.session_state["custom_floors"]:
                with st.container():
                    f_ren_col1, f_ren_col2, f_ren_col3 = st.columns([1.5, 2, 1])
                    with f_ren_col1:
                        target_floor = st.selectbox("Select to Rename:", st.session_state["custom_floors"], key="sb_target_floor", label_visibility="collapsed")
                    with f_ren_col2:
                        renamed_floor = st.text_input("New Name:", placeholder="New floor name", key="txt_rename_floor", label_visibility="collapsed")
                    with f_ren_col3:
                        if st.button("Rename", key="btn_rename_floor", use_container_width=True) and renamed_floor.strip():
                            f_new_clean = renamed_floor.strip()
                            idx = st.session_state["custom_floors"].index(target_floor)
                            st.session_state["custom_floors"][idx] = f_new_clean
                            st.rerun()

            # ফ্লোর ডিলিট করার ডেডিকেটেড অপশন (Delete Section)
            if st.session_state["custom_floors"]:
                with st.container():
                    f_del_col1, f_del_col2 = st.columns([3.5, 1])
                    with f_del_col1:
                        delete_target_floor = st.selectbox("Select to Delete:", st.session_state["custom_floors"], key="sb_delete_floor", label_visibility="collapsed")
                    with f_del_col2:
                        if st.button("Delete", key="btn_delete_floor", use_container_width=True):
                            st.session_state["custom_floors"].remove(delete_target_floor)
                            st.rerun()

            st.markdown("**Active Floors (Check to activate):**")
            selected_floors = []
            
            # 🚀 ৫ কলাম বিশিষ্ট কাস্টম স্পেস-সেভিং গ্রিড লেআউট (লাইন স্পেস মিনিমাইজড)
            if st.session_state["custom_floors"]:
                floor_list = list(st.session_state["custom_floors"])
                cols_per_row = 5
                for i in range(0, len(floor_list), cols_per_row):
                    row_floors = floor_list[i:i+cols_per_row]
                    grid_cols = st.columns(cols_per_row)
                    for idx, f in enumerate(row_floors):
                        with grid_cols[idx]:
                            if st.checkbox(f, value=True, key=f"chk_{f}"):
                                selected_floors.append(f)
                        
        st.markdown("---")
        if st.button("🔀 Execute Random Floor & Room Assignment", use_container_width=True, type="primary"):
            if not selected_rooms or not selected_floors:
                st.error("❌ Please select at least one Room and one Floor before execution.")
            else:
                run_random_assignment(df, selected_rooms, selected_floors)

    st.markdown("Filter the duty roster below. **Leave a filter empty to include all values.**")

    f1, f2, f3, f4 = st.columns(4)
    with f1:
        sel_dates = st.multiselect("📅 Date", sorted(df["Date"].unique()))
    with f2:
        sel_slots = st.multiselect("⏰ Time Slot", sorted(df["Time Slot"].unique()))
    with f3:
        sel_desigs = st.multiselect("🎓 Designation", sorted(df["Designation"].unique()))
    with f4:
        sel_cats = st.multiselect("🏷️ Category", sorted(df["Role"].unique()))

    filtered = df
    if sel_dates:
        filtered = filtered[filtered["Date"].isin(sel_dates)]
    if sel_slots:
        filtered = filtered[filtered["Time Slot"].isin(sel_slots)]
    if sel_desigs:
        filtered = filtered[filtered["Designation"].isin(sel_desigs)]
    if sel_cats:
        filtered = filtered[filtered["Role"].isin(sel_cats)]

    # ম্যাপিং প্রসেস: মেইন টেবিলে সুনির্দিষ্ট ডাটা পুশ করা
    assigned_column = []
    for _, row in filtered.iterrows():
        s_key = f"{row['Date']}_{row['Time Slot']}"
        p_name = row['Name']
        assigned_val = st.session_state["assignments"].get(s_key, {}).get(p_name, "Not Assigned Yet")
        assigned_column.append(assigned_val)
        
    filtered = filtered.copy()
    filtered["Floor / Room Assignment"] = assigned_column

    personnel = (
        filtered[["Name", "Designation", "Role", "Technology/Dept", "Date", "Time Slot", "Duty Status", "Floor / Room Assignment"]]
        .rename(columns={"Role": "Category", "Technology/Dept": "Dept/Tech"})
        .sort_values(["Date", "Time Slot", "Name"])
        .reset_index(drop=True)
    )
    personnel.index = personnel.index + 1

    m1, m2, m3 = st.columns(3)
    m1.metric("Duty Assignments", len(personnel))
    m2.metric("Unique Personnel", personnel["Name"].nunique())
    m3.metric("Dates Covered", personnel["Date"].nunique())

    st.markdown("#### 👥 Personnel on Duty")
    st.markdown('<div id="personnel-on-duty"></div>', unsafe_allow_html=True)
    if personnel.empty:
        st.info("No personnel match the selected filters.")
    else:
        # স্ক্রিন ডিসপ্লের কলাম লেবেল ডাইনামিক করা (ক্যাটাগরি ওয়াইজ)
        display_label = "Floor / Room Assignment"
        if sel_cats and len(sel_cats) == 1:
            cat_selected = sel_cats[0]
            if "Invigilator" in cat_selected:
                display_label = "Room No"
            elif "MLSS" in cat_selected or "Hall Super" in cat_selected:
                display_label = "Floor No"

        screen_df = personnel.rename(columns={"Floor / Room Assignment": display_label})
        st.dataframe(screen_df, use_container_width=True)
        
        # --- আপনার লেআউট অনুযায়ী কাস্টম এক্সেল জেনারেটর ইঞ্জিন ---
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Duty Roster"
        ws.views.sheetView[0].showGridLines = True # গ্রিডলাইন অন রাখা
        
        # ১. প্রাতিষ্ঠানিক হেডার (Row 1 & 2)
        ws.merge_cells("A1:F1")
        ws["A1"] = "Narsingdi Government Polytechnic Institute"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells("A2:F2")
        ws["A2"] = "Mid-Term Examination - 2026 | Personalized Daily Duty Roster"
        ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="475569")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        
        # ২. মেটাডেটা ইনফরমেশন (Row 4) - ফিল্টারের ওপর ভিত্তি করে ডাইনামিক হবে
        meta_cat = sel_cats[0] if sel_cats else "All Categories"
        meta_date = sel_dates[0] if sel_dates else "All Dates"
        meta_slot = sel_slots[0] if sel_slots else "All Times"
        
        ws["B4"] = f"Category: {meta_cat}"
        ws["D4"] = f"Date: {meta_date}"
        ws["F4"] = f"Time: {meta_slot}"
        for col in ["B", "D", "F"]:
            ws[f"{col}4"].font = Font(name="Calibri", size=10, bold=True)
            
        # ৩. ডাইনামিক কলাম হেডার নির্ধারণ
        dynamic_assignment_label = "Floor / Room Assignment"
        if sel_cats and len(sel_cats) == 1:
            if "Invigilator" in sel_cats[0]:
                dynamic_assignment_label = "Room No"
            elif "MLSS" in sel_cats[0] or "Hall Super" in sel_cats[0]:
                dynamic_assignment_label = "Floor No"

        headers = ["SL", "Name", "Designation", "Dept/Tech", dynamic_assignment_label, "Signature"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header_title
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # ৪. ডেটা পপুলেশন (Row 7 থেকে শুরু)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        current_row = 7
        for idx, row in personnel.iterrows():
            ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=row["Name"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=3, value=row["Designation"]).alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=4, value=row["Dept/Tech"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=5, value=row["Floor / Room Assignment"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=6, value="").alignment = Alignment(horizontal="center") # ফিজিক্যাল সিগনেচার স্পেস খালি রাখা
            
            # বর্ডার এবং ফন্ট স্টাইল সেট করা
            for col_num in range(1, 7):
                c = ws.cell(row=current_row, column=col_num)
                c.font = Font(name="Calibri", size=11)
                c.border = thin_border
                # অল্টারনেティブ রো কালারিং (হালকা গ্রে শেড)
                if current_row % 2 == 0:
                    c.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            current_row += 1
            
        # ৫. প্রাতিষ্ঠানিক সমাপ্তি / সিগনেচার এরিয়া (Row + 3)
        current_row += 3
        ws.cell(row=current_row, column=6, value="Control Room Officer").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row+1, column=6, value="Narsingdi GPI").font = Font(name="Calibri", size=10, italic=True)
        ws.cell(row=current_row+1, column=6).alignment = Alignment(horizontal="center")
        
        # কলামের উইডথ স্বয়ংক্রিয়ভাবে অ্যাডজাস্ট করা (Auto-fit Columns)
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
        ws.column_dimensions['F'].width = 16 # সিগনেচার স্পেস একটু বড় রাখা
            
        # ফাইলটি বাফারে সেভ করা
        buffer_excel = io.BytesIO()
        wb.save(buffer_excel)
        excel_data = buffer_excel.getvalue()
        
        # ৬. ডাইনামিক ফাইলের নাম জেনারেশন ইঞ্জিন (Date_Time_Category)
        fn_date = str(sel_dates[0]).replace("/", "-") if sel_dates else "All-Dates"
        fn_time = str(sel_slots[0]).replace(" ", "-").replace(":", "-") if sel_slots else "All-Times"
        fn_cat = str(sel_cats[0]).replace(" ", "-").replace("/", "-") if sel_cats else "All-Categories"
        custom_filename = f"Roster_{fn_date}_{fn_time}_{fn_cat}.xlsx"
        
        # 📥 রেডি-মেড লেআউট এক্সেল ডাউনলোড বাটন
        st.download_button(
            label="📥 Download Ready-Made Official Roster (Excel)",
            data=excel_data,
            file_name=custom_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# ৪. রস্টার ডেটা লোড
try:
    df_portal = load_and_process_roster()
except Exception as e:
    st.error(f"এক্সেল ফাইলটি লোড করা যাচ্ছে না। ত্রুটি: {e}")
    st.stop()

# ৫. পোর্টাল টপ ব্যানার রেন্ডারিং
# Language selector buttons - compact
col_en, col_bn = st.columns(2, gap="small")
with col_en:
    if st.button("🇬🇧 EN", use_container_width=True, key="btn_en"):
        st.session_state["language"] = "en"
        st.rerun()
with col_bn:
    if st.button("🇧🇩 BN", use_container_width=True, key="btn_bn"):
        st.session_state["language"] = "bn"
        st.rerun()

logo_element = f"<img src='{logo_b64}' class='header-img-render'>" if logo_b64 else "<div class='fallback-logo-render'>📝</div>"
st.markdown(f"""
<div class='portal-header-container'>
    {logo_element}
    <div class='header-text'>
        <h1>{"Narsingdi Government Polytechnic Institute" if st.session_state["language"] == "en" else "নরসিংদী সরকারি পলিটেকনিক ইনস্টিটিউট"}</h1>
        <p>পর্বমধ্য পরীক্ষা - ২০২৬ | ডিজিটাল ডিউটি রোস্টার ম্যানেজমেন্ট পোর্টাল</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ৫.৫ নেভিগেশন — মূল পেজে ভিউ সুইচার
app_view = st.segmented_control(
    "Navigation",
    ["🏠 Personal Roster", "🛡️ Control Room"],
    default="🏠 Personal Roster",
    label_visibility="collapsed",
)

if app_view == "🛡️ Control Room":
    render_control_room(df_portal)
    st.stop()

# 六. ইন-পেজ ফিল্টার প্যানেল
st.markdown("<div class='filter-section-title'>🎛️ নিচের অপশনগুলো থেকে আপনার নাম সিলেক্ট করুন:</div>", unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)

with col1:
    selected_role = st.selectbox("১. পদবি/ক্যাটাগরি নির্ধারণ করুন:", sorted(df_portal['Role'].unique()), index=None, placeholder="— নির্বাচন করুন —", key="main_role")

df_filtered_role = df_portal[df_portal['Role'] == selected_role] if selected_role else df_portal.iloc[0:0]

with col2:
    selected_tech = st.selectbox("২. টেকনোলজি/ডিপার্টমেন্ট নির্বাচন করুন:", sorted(df_filtered_role['Technology/Dept'].unique()), index=None, placeholder="— নির্বাচন করুন —", key="main_tech")

df_filtered_tech = df_filtered_role[df_filtered_role['Technology/Dept'] == selected_tech] if selected_tech else df_portal.iloc[0:0]

with col3:
    selected_name = st.selectbox("৩. তালিকায় আপনার নাম নির্বাচন করুন:", sorted(df_filtered_tech['Name'].unique()), index=None, placeholder="— নির্বাচন করুন —", key="main_name")

# ৭. ডেটা ফিল্টার ও ভিউ জেনারেশন
if not (selected_role and selected_tech and selected_name):
    st.info("👆 আপনার ডিউটি রোস্টার দেখতে উপরের তিনটি অপশন (পদবি, টেকনোলজি/ডিপার্টমেন্ট ও নাম) নির্বাচন করুন।")
    st.stop()

# ইউজার ইন্টারফেসেও এসাইনমেন্ট সিঙ্ক করা
df_mapped = df_filtered_tech.copy()
assigned_column_user = []
for _, row in df_mapped.iterrows():
    s_key = f"{row['Date']}_{row['Time Slot']}"
    p_name = row['Name']
    assigned_val = st.session_state["assignments"].get(s_key, {}).get(p_name, "Not Assigned Yet")
    assigned_column_user.append(assigned_val)
df_mapped["Floor / Room Assignment"] = assigned_column_user

user_schedule = df_mapped[df_mapped['Name'] == selected_name][['Date', 'Time Slot', 'Duty Status', 'Floor / Room Assignment']]
user_info = df_mapped[df_mapped['Name'] == selected_name].iloc[0]

# প্রোফাইল কার্ড প্রদর্শন
st.markdown(f"""
<div class='profile-card'>
    <div class='profile-grid'>
        <div class='profile-item'><strong>Name:</strong> {user_info['Name']}</div>
        <div class='profile-item'><strong>Dept/Tech:</strong> {user_info['Technology/Dept']}</div>
        <div class='profile-item'><strong>Designation:</strong> {user_info['Designation']}</div>
        <div class='profile-item'><strong>Category:</strong> {user_info['Role']}</div>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("<h4 style='color: #1E3A8A; margin-bottom: 15px;'>🗓 Your Personalized Duty Roster:</h4>", unsafe_allow_html=True)

if not user_schedule.empty:
    user_schedule.reset_index(drop=True, inplace=True)
    user_schedule.index = user_schedule.index + 1
    
    st.dataframe(user_schedule, use_container_width=True)
    
    # ৮. পিডিএফ জেনারেশন ও ডাউনলোড বাটন
    pdf_bytes = generate_pdf(user_info, user_schedule)
    pdf_filename = f"Duty_Roster_{str(user_info['Name']).replace(' ', '_')}.pdf"

    st.download_button(
        label="📥 Download PDF Document",
        data=pdf_bytes,
        file_name=pdf_filename,
        mime="application/pdf",
    )
else:
    st.success("🎉 No exam duty assigned to your profile in this roster.")